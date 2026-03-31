"""One-time migration script: Excel -> SQLite."""
import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from db.database import init_db, get_connection, query
from db.seed_data import seed_all

EXCEL_PATH = Path(__file__).parent.parent / "11150-Emergent Needs Report Tracker.xlsm"

# Map role names from Excel to job_codes titles
ROLE_MAP = {
    "Analyst 2 - Imagery": "Analyst 2 - Imagery",
    "Analyst 1 - Imagery": "Analyst 1 - Imagery",
    "Analyst 3 - Imagery": "Analyst 3 - Imagery",
    "Analyst 2 - GeoINT": "Analyst 2 - GeoINT",
    "Analyst 1 - GeoINT": "Analyst 1 - GeoINT",
    "Analyst 3 - GeoINT": "Analyst 3 - GeoINT",
    "Data Scientist 2": "Data Scientist 2",
    "Data Manager 1": "Data Manager 1",
    "OSINT 2": "OSINT 2",
    "Contract PAI": "Contract PAI",
}


def import_project(ws, sheet_name, conn, job_code_lookup):
    """Import a single report sheet as a project."""
    # Extract header info
    pws = str(ws.cell(1, 2).value or "")
    # Title can be in E1 or D1 depending on the sheet format
    title = ws.cell(1, 5).value or ws.cell(1, 4).value or sheet_name
    start_val = ws.cell(2, 2).value
    end_val = ws.cell(3, 2).value
    status = ws.cell(2, 5).value or ws.cell(2, 4).value or "Ongoing"
    notes = ws.cell(3, 5).value or ws.cell(3, 4).value or ""
    days = ws.cell(4, 2).value or 1

    # Format dates
    start_date = None
    end_date = None
    if isinstance(start_val, datetime):
        start_date = start_val.strftime("%Y-%m-%d")
    if isinstance(end_val, datetime):
        end_date = end_val.strftime("%Y-%m-%d")

    # Check N5 for daily rate flag
    is_daily_rate = 1 if ws.cell(5, 14).value is True else 0

    # Insert project
    cur = conn.execute(
        """INSERT INTO projects (pws_number, report_title, start_date, end_date, status, notes, days, is_daily_rate)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (pws, str(title), start_date, end_date, str(status), str(notes) if notes else None, days, is_daily_rate),
    )
    project_id = cur.lastrowid
    print(f"  Imported project: {title} (PWS: {pws}, ID: {project_id})")

    # Import labor entries (rows 7+ until empty or "Total")
    labor_count = 0
    for row in range(7, 20):
        role = ws.cell(row, 1).value
        if role is None or str(role).strip() == "" or "Total" in str(role):
            break

        role_str = str(role).strip()
        hours = ws.cell(row, 3).value or 0
        emp_rate = ws.cell(row, 2).value or 0
        bid_rate = ws.cell(row, 5).value or 0

        # Look up job code
        jc_id = job_code_lookup.get(role_str)
        if jc_id is None:
            # Try partial match
            for title_key, jc_data in job_code_lookup.items():
                if role_str.lower() in title_key.lower() or title_key.lower() in role_str.lower():
                    jc_id = jc_data
                    break
        if jc_id is None:
            print(f"    WARNING: Unknown role '{role_str}', skipping")
            continue

        conn.execute(
            """INSERT INTO labor_entries (project_id, job_code_id, person_name, hours, employee_rate, bid_rate)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (project_id, jc_id, None, float(hours), float(emp_rate), float(bid_rate)),
        )
        labor_count += 1

    print(f"    {labor_count} labor entries")

    # Import imagery orders - find the imagery header row dynamically
    img_header_row = None
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row, 1).value
        if cell_val and "Imagery Provider" in str(cell_val):
            img_header_row = row
            break

    img_count = 0
    if img_header_row:
        for row in range(img_header_row + 1, ws.max_row + 1):
            provider = ws.cell(row, 1).value
            if provider is None or "Imagery Total" in str(provider):
                break

            provider_str = str(provider).strip()
            if not provider_str:
                continue

            # Column layout varies - handle both formats
            product = ws.cell(row, 2).value or ""
            order_date_val = ws.cell(row, 3).value
            delivered_val = ws.cell(row, 4).value
            aoi = ws.cell(row, 5).value or ""
            shots = ws.cell(row, 6).value or 0
            cost = ws.cell(row, 7).value or 0
            charge = ws.cell(row, 8).value or 0

            order_date = None
            if isinstance(order_date_val, datetime):
                order_date = order_date_val.strftime("%Y-%m-%d")
            elif order_date_val:
                order_date = str(order_date_val)

            delivered = 0
            if delivered_val:
                delivered = 1 if str(delivered_val).strip().upper() in ("Y", "YES", "TRUE", "1") else 0

            conn.execute(
                """INSERT INTO imagery_orders (project_id, catalog_id, provider, product, order_date,
                   delivered, aoi, shots_delivered, cost, charge)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (project_id, None, provider_str, str(product), order_date,
                 delivered, str(aoi) if aoi else None, int(shots) if shots else 0,
                 float(cost) if cost else 0, float(charge) if charge else 0),
            )
            img_count += 1

    print(f"    {img_count} imagery orders")


def main():
    print(f"Importing from: {EXCEL_PATH}")

    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        return

    # Initialize DB and seed reference data
    init_db()
    seed_all()

    # Build job code lookup
    job_codes = query("SELECT * FROM job_codes")
    job_code_lookup = {jc["title"]: jc["id"] for jc in job_codes}

    # Open workbook
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True, keep_vba=True)
    print(f"Sheets: {wb.sheetnames}")

    conn = get_connection()

    # Skip these sheets - they're not project reports
    skip_sheets = {"Home", "Job_Codes", "Imagery Pricing 2", "Report_Template",
                   "TacSRT_Tracker_Condensed", "Running Profit"}

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            print(f"Skipping: {sheet_name}")
            continue

        ws = wb[sheet_name]
        # Verify this looks like a report sheet (has PWS in A1)
        if ws.cell(1, 1).value == "PWS":
            print(f"\nImporting sheet: {sheet_name}")
            import_project(ws, sheet_name, conn, job_code_lookup)
        else:
            print(f"Skipping (not a report sheet): {sheet_name}")

    conn.commit()
    conn.close()
    print("\nImport complete!")


if __name__ == "__main__":
    main()
